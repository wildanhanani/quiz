import io

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from .models import Category, Choice, Question, QuizAttempt, Subscription


class QuizAccessTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username='basicuser',
            password='StrongPass123!',
            is_active=True,
        )
        self.subscription = self.user.subscription
        self.subscription.package = 'BASIC'
        self.subscription.max_categories = 5
        self.subscription.max_attempts_per_quiz = 2
        self.subscription.save()

        self.categories = [
            Category.objects.create(name=f'Category {index}', description='Test category')
            for index in range(1, 7)
        ]
        self.allowed_category = self.categories[0]
        self.blocked_category = self.categories[5]
        for index in range(1, 4):
            question = Question.objects.create(
                category=self.allowed_category,
                text=f'Soal latihan {index}?',
                order=index,
            )
            Choice.objects.create(question=question, text='Benar', is_correct=True)
            Choice.objects.create(question=question, text='Salah', is_correct=False)

        self.client.login(username='basicuser', password='StrongPass123!')

    def test_dashboard_only_shows_first_five_categories_for_basic_package(self):
        response = self.client.get(reverse('dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Category 1')
        self.assertContains(response, 'Category 5')
        self.assertNotContains(response, 'Category 6')

    def test_dashboard_groups_subcategories_under_parent_category(self):
        parent = Category.objects.create(name='CPNS', description='Kelompok soal CPNS')
        self.allowed_category.parent = parent
        self.allowed_category.save(update_fields=['parent'])
        self.categories[1].parent = parent
        self.categories[1].save(update_fields=['parent'])
        self.subscription.max_categories = 6
        self.subscription.save(update_fields=['max_categories'])

        response = self.client.get(reverse('dashboard'))

        self.assertContains(response, 'CPNS')
        self.assertContains(response, 'Category 1')
        self.assertContains(response, 'Category 2')
        self.assertContains(response, 'Subkategori Mandiri')

    def test_direct_url_to_blocked_category_is_rejected_for_basic_package(self):
        response = self.client.get(reverse('take_quiz', args=[self.blocked_category.id]))

        self.assertRedirects(response, reverse('dashboard'))

    def test_take_quiz_page_allows_user_to_configure_timer(self):
        response = self.client.get(reverse('take_quiz', args=[self.allowed_category.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Timer')
        self.assertContains(response, '1 soal = 1 menit')
        self.assertContains(response, '3 menit total untuk 3 soal.')
        self.assertContains(response, 'Kurang dari 5 menit.')
        self.assertContains(response, 'Kurang dari 1 menit.')
        self.assertContains(response, 'Jawaban tersimpan otomatis di browser ini.')
        self.assertContains(response, 'quiz-autosave-', html=False)
        self.assertContains(response, 'Progress')
        self.assertContains(response, 'soal terjawab')
        self.assertContains(response, 'beforeunload', html=False)
        self.assertContains(response, 'Belum dijawab')
        self.assertContains(response, 'unansweredList', html=False)
        self.assertContains(response, 'Pagination Soal')
        self.assertContains(response, 'miniNavigator', html=False)
        self.assertContains(response, 'Tampilkan satu soal per halaman')
        self.assertContains(response, 'paginationAnsweredCount', html=False)
        self.assertContains(response, '0 dari 3 soal terjawab')
        self.assertContains(response, 'previousQuestionButton', html=False)
        self.assertContains(response, 'nextQuestionButton', html=False)
        self.assertContains(response, 'Review Sebelum Submit')
        self.assertContains(response, 'confirmSubmitQuiz', html=False)
        self.assertNotContains(response, 'required', html=False)
        self.assertNotContains(response, 'Gunakan timer')
        self.assertNotContains(response, 'Durasi timer')
        self.assertNotContains(response, 'Preset cepat')

    def test_take_quiz_page_uses_automatic_timer_based_on_total_questions(self):
        self.subscription.prefers_timer = True
        self.subscription.preferred_timer_minutes = 25
        self.subscription.save()

        response = self.client.get(reverse('take_quiz', args=[self.allowed_category.id]))

        self.assertContains(response, 'Timer aktif 3 menit')
        self.assertContains(response, '3 menit total untuk 3 soal.')
        self.assertNotContains(response, '25 menit')

    def test_dashboard_displays_category_analytics_summary(self):
        QuizAttempt.objects.create(
            user=self.user,
            category=self.allowed_category,
            score=3,
            total_questions=5,
            answers={},
        )
        QuizAttempt.objects.create(
            user=self.user,
            category=self.allowed_category,
            score=4,
            total_questions=5,
            answers={},
        )

        response = self.client.get(reverse('dashboard'))

        self.assertContains(response, 'Total Attempt')
        self.assertContains(response, '2')
        self.assertContains(response, 'Rata-rata Akurasi')
        self.assertContains(response, '70%')
        self.assertContains(response, 'Category 1')
        self.assertContains(response, '80%')
        self.assertContains(response, 'Naik 20%')


class QuizAttemptLimitTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username='freeuser',
            password='StrongPass123!',
            is_active=True,
        )
        self.subscription = self.user.subscription
        self.subscription.package = 'FREE'
        self.subscription.max_attempts_per_quiz = 1
        self.subscription.save()

        self.category = Category.objects.create(
            name='Free Category',
            description='Accessible category',
            is_premium=False,
        )
        self.question = Question.objects.create(
            category=self.category,
            text='2 + 2 = ?',
            order=1,
        )
        self.correct_choice = Choice.objects.create(
            question=self.question,
            text='4',
            is_correct=True,
        )
        Choice.objects.create(question=self.question, text='5', is_correct=False)

        self.client.login(username='freeuser', password='StrongPass123!')

    def test_attempt_submission_creates_result(self):
        response = self.client.post(
            reverse('take_quiz', args=[self.category.id]),
            {
                f'question_{self.question.id}': str(self.correct_choice.id),
                'use_timer': 'on',
                'timer_minutes': '15',
            },
        )

        attempt = QuizAttempt.objects.get(user=self.user, category=self.category)
        self.subscription.refresh_from_db()
        self.assertRedirects(response, reverse('quiz_result', args=[attempt.id]))
        self.assertEqual(attempt.score, 1)
        self.assertEqual(attempt.total_questions, 1)
        self.assertTrue(attempt.answers['_meta']['timer_enabled'])
        self.assertEqual(attempt.answers['_meta']['timer_minutes'], 1)

    def test_take_quiz_recreates_missing_subscription_automatically(self):
        self.subscription.delete()

        response = self.client.get(reverse('take_quiz', args=[self.category.id]))

        self.assertEqual(response.status_code, 200)
        recreated_subscription = Subscription.objects.get(user=self.user)
        self.assertEqual(recreated_subscription.package, 'FREE')

    def test_subscription_ensure_for_user_reuses_existing_record(self):
        ensured_subscription = Subscription.ensure_for_user(self.user)

        self.assertEqual(ensured_subscription.id, self.subscription.id)
        self.assertEqual(Subscription.objects.filter(user=self.user).count(), 1)

    def test_user_cannot_retry_after_reaching_attempt_limit(self):
        QuizAttempt.objects.create(
            user=self.user,
            category=self.category,
            score=1,
            total_questions=1,
            answers={str(self.question.id): {'selected': self.correct_choice.id, 'correct': True}},
        )

        response = self.client.get(reverse('take_quiz', args=[self.category.id]))

        self.assertRedirects(response, reverse('dashboard'))
        self.assertEqual(
            QuizAttempt.objects.filter(user=self.user, category=self.category).count(),
            1,
        )

    def test_dashboard_hides_start_cta_when_attempt_limit_is_reached(self):
        QuizAttempt.objects.create(
            user=self.user,
            category=self.category,
            score=1,
            total_questions=1,
            answers={},
        )

        response = self.client.get(reverse('dashboard'))

        self.assertContains(response, 'Batas attempt habis')
        self.assertNotContains(response, reverse('take_quiz', args=[self.category.id]))

    def test_free_user_cannot_access_subcategory_under_premium_parent(self):
        premium_parent = Category.objects.create(
            name='CPNS Premium',
            description='Premium package',
            is_premium=True,
        )
        self.category.parent = premium_parent
        self.category.save(update_fields=['parent'])

        response = self.client.get(reverse('take_quiz', args=[self.category.id]))

        self.assertRedirects(response, reverse('dashboard'))

    def test_result_page_displays_attempt_number_and_personal_best(self):
        QuizAttempt.objects.create(
            user=self.user,
            category=self.category,
            score=0,
            total_questions=1,
            answers={
                '_meta': {'timer_enabled': False, 'timer_minutes': None},
                str(self.question.id): {'selected': self.correct_choice.id, 'correct': False},
            },
        )
        latest_attempt = QuizAttempt.objects.create(
            user=self.user,
            category=self.category,
            score=1,
            total_questions=1,
            answers={
                '_meta': {'timer_enabled': True, 'timer_minutes': 12},
                str(self.question.id): {'selected': self.correct_choice.id, 'correct': True},
            },
        )

        response = self.client.get(reverse('quiz_result', args=[latest_attempt.id]))

        self.assertContains(response, 'Attempt Kategori Ini')
        self.assertContains(response, '2')
        self.assertContains(response, 'Personal Best')
        self.assertContains(response, '100%')
        self.assertContains(response, 'Mode Pengerjaan')
        self.assertContains(response, 'Timer 12 menit')

    def test_result_page_hides_retry_cta_when_attempt_limit_is_reached(self):
        attempt = QuizAttempt.objects.create(
            user=self.user,
            category=self.category,
            score=1,
            total_questions=1,
            answers={
                '_meta': {'timer_enabled': False, 'timer_minutes': None},
                str(self.question.id): {'selected': self.correct_choice.id, 'correct': True},
            },
        )

        response = self.client.get(reverse('quiz_result', args=[attempt.id]))

        self.assertContains(response, 'Batas attempt habis')
        self.assertNotContains(response, 'Ulangi Quiz')

    def test_attempt_submission_ignores_choice_from_another_question(self):
        second_question = Question.objects.create(
            category=self.category,
            text='3 + 3 = ?',
            order=2,
        )
        second_question_correct_choice = Choice.objects.create(
            question=second_question,
            text='6',
            is_correct=True,
        )
        Choice.objects.create(question=second_question, text='7', is_correct=False)

        response = self.client.post(
            reverse('take_quiz', args=[self.category.id]),
            {f'question_{self.question.id}': str(second_question_correct_choice.id)},
        )

        attempt = QuizAttempt.objects.get(user=self.user, category=self.category)
        self.assertRedirects(response, reverse('quiz_result', args=[attempt.id]))
        self.assertEqual(attempt.score, 0)
        self.assertNotIn(str(self.question.id), attempt.answers)


class AdminFeaturesTests(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            username='admin_test',
            email='admin@example.com',
            password='StrongPass123!',
        )
        self.client.login(username='admin_test', password='StrongPass123!')

    def test_admin_dashboard_shows_statistics_cards(self):
        user = User.objects.create_user(
            username='member',
            password='StrongPass123!',
            is_active=True,
        )
        category = Category.objects.create(name='Admin Stats Category', description='Stats test')
        question = Question.objects.create(category=category, text='Question stats?', order=1)
        Choice.objects.create(question=question, text='A', is_correct=True)
        Choice.objects.create(question=question, text='B', is_correct=False)
        QuizAttempt.objects.create(
            user=user,
            category=category,
            score=1,
            total_questions=1,
            answers={},
        )

        response = self.client.get(reverse('admin:index'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Admin Dashboard')
        self.assertContains(response, 'Total User')
        self.assertContains(response, 'Total Attempt')
        self.assertContains(response, 'Kategori Paling Sering Dikerjakan')
        self.assertContains(response, 'Attempt Terbaru')
        self.assertContains(response, 'Import Soal CSV / XLSX / ZIP')
        self.assertContains(response, 'Download Template XLSX')

    def test_admin_can_download_xlsx_import_template(self):
        response = self.client.get(reverse('admin:export_questions_template'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('question-import-template.xlsx', response['Content-Disposition'])

        workbook = load_workbook(io.BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ['Questions Template', 'Instructions'])

        template_sheet = workbook['Questions Template']
        self.assertEqual(
            [cell.value for cell in template_sheet[1]],
            [
                'parent_category',
                'category',
                'category_description',
                'question',
                'explanation',
                'image',
                'order',
                'choice_1',
                'choice_2',
                'choice_3',
                'choice_4',
                'choice_5',
                'correct_answer',
            ],
        )
        self.assertEqual(template_sheet['A2'].value, 'CPNS')
        self.assertEqual(template_sheet['B2'].value, 'TWK')
        self.assertEqual(template_sheet['D2'].value, 'Lambang sila pertama Pancasila adalah?')
        self.assertIn('bintang', template_sheet['E2'].value.lower())

        instructions_sheet = workbook['Instructions']
        self.assertEqual(instructions_sheet['A2'].value, 'parent_category')
        self.assertIn('kategori utama', instructions_sheet['B2'].value.lower())
        self.assertEqual(instructions_sheet['A6'].value, 'explanation')
        workbook.close()

    def test_admin_can_import_questions_from_xlsx(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            'parent_category',
            'category',
            'category_description',
            'question',
            'explanation',
            'image',
            'order',
            'choice_1',
            'choice_2',
            'choice_3',
            'choice_4',
            'choice_5',
            'correct_answer',
        ])
        sheet.append([
            'CPNS',
            'XLSX Category',
            'Imported from workbook',
            'Apa hasil 9 + 1?',
            'Penjumlahan sederhana menghasilkan angka 10.',
            '',
            1,
            '8',
            '9',
            '10',
            '11',
            '',
            3,
        ])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        upload = SimpleUploadedFile(
            'questions.xlsx',
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response = self.client.post(
            reverse('admin:import_questions'),
            {'csv_file': upload},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(Question.objects.filter(text='Apa hasil 9 + 1?').exists())
        imported_question = Question.objects.get(text='Apa hasil 9 + 1?')
        self.assertEqual(imported_question.category.name, 'XLSX Category')
        self.assertEqual(imported_question.category.parent.name, 'CPNS')
        self.assertEqual(imported_question.explanation, 'Penjumlahan sederhana menghasilkan angka 10.')
        self.assertEqual(imported_question.choices.count(), 4)
        self.assertEqual(imported_question.choices.get(is_correct=True).text, '10')
        self.assertContains(response, 'Berhasil mengimpor 1 soal dari Excel (.xlsx)!')

    def test_admin_import_rolls_back_entire_file_when_one_row_is_invalid(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            'parent_category',
            'category',
            'category_description',
            'question',
            'explanation',
            'image',
            'order',
            'choice_1',
            'choice_2',
            'choice_3',
            'choice_4',
            'choice_5',
            'correct_answer',
        ])
        sheet.append([
            'CPNS',
            'Rollback Category',
            'Imported from workbook',
            'Apa hasil 2 + 2?',
            'Penjumlahan sederhana menghasilkan angka 4.',
            '',
            1,
            '3',
            '4',
            '',
            '',
            '',
            2,
        ])
        sheet.append([
            'CPNS',
            'Rollback Category',
            'Imported from workbook',
            'Baris invalid',
            'Pembahasan invalid.',
            '',
            2,
            'Pilihan tunggal',
            '',
            '',
            '',
            '',
            1,
        ])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        upload = SimpleUploadedFile(
            'questions-invalid.xlsx',
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response = self.client.post(
            reverse('admin:import_questions'),
            {'csv_file': upload},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Question.objects.filter(category__name='Rollback Category').count(), 0)
        self.assertContains(response, 'Baris 3')
        self.assertContains(response, 'minimal harus ada 2 pilihan jawaban')

    def test_admin_import_rejects_conflicting_existing_category_structure(self):
        existing_category = Category.objects.create(name='TWK', description='Kategori lama')

        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            'parent_category',
            'category',
            'category_description',
            'question',
            'explanation',
            'image',
            'order',
            'choice_1',
            'choice_2',
            'choice_3',
            'choice_4',
            'choice_5',
            'correct_answer',
        ])
        sheet.append([
            'CPNS',
            'TWK',
            'Imported from workbook',
            'Soal baru',
            'Pembahasan baru.',
            '',
            1,
            'A',
            'B',
            '',
            '',
            '',
            1,
        ])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        upload = SimpleUploadedFile(
            'questions-conflict.xlsx',
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response = self.client.post(
            reverse('admin:import_questions'),
            {'csv_file': upload},
            follow=True,
        )

        existing_category.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertIsNone(existing_category.parent)
        self.assertEqual(Question.objects.filter(text='Soal baru').count(), 0)
        self.assertContains(response, 'sudah ada pada struktur lain')
