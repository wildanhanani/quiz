from django.contrib import admin
from django import forms
from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from django.urls import path
from django.contrib import messages
from django.db import transaction
from django.db.models import Count, Sum
from django.http import HttpResponse
from django.utils import timezone
from django.utils.html import format_html
from .models import Category, Question, Choice, QuizAttempt, Subscription
import csv
import os
import shutil
import tempfile
import zipfile
from io import BytesIO

from django.core.files import File
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def normalize_row(row):
    return {str(key).strip(): str(value).strip() if value is not None else '' for key, value in row.items()}


def iter_csv_rows(file_path):
    with open(file_path, 'r', encoding='utf-8-sig') as handle:
        sample = handle.read(2048)
        handle.seek(0)
        first_line = sample.split('\n')[0]
        delimiter = ';' if first_line.count(';') > first_line.count(',') else ','
        reader = csv.DictReader(handle, delimiter=delimiter)
        reader.fieldnames = [name.strip() for name in reader.fieldnames]
        rows = [normalize_row(row) for row in reader]
    return rows, f'CSV (delimiter: {delimiter})'


def iter_xlsx_rows(file_path):
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    sheet = workbook.active
    raw_rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not raw_rows:
        return [], 'Excel (.xlsx)'

    headers = [str(value).strip() if value is not None else '' for value in raw_rows[0]]
    rows = []
    for values in raw_rows[1:]:
        row = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = values[index] if index < len(values) else ''
            row[header] = value
        rows.append(normalize_row(row))
    return rows, 'Excel (.xlsx)'


def load_question_rows(file_path):
    extension = os.path.splitext(file_path)[1].lower()
    if extension == '.csv':
        return iter_csv_rows(file_path)
    if extension == '.xlsx':
        return iter_xlsx_rows(file_path)
    raise ValueError('Supported question files are .csv and .xlsx')


QUESTION_TEMPLATE_HEADERS = [
    'parent_category',
    'category',
    'category_description',
    'question',
    'explanation',
    'image',
    'order',
    'choice_1',
    'choice_2',
    'choice_3',
    'choice_4',
    'choice_5',
    'correct_answer',
]


def build_question_template_workbook():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Questions Template'
    sheet.append(QUESTION_TEMPLATE_HEADERS)
    sheet.append([
        'CPNS',
        'TWK',
        'Tes Wawasan Kebangsaan',
        'Lambang sila pertama Pancasila adalah?',
        'Sila pertama dilambangkan dengan bintang sebagai simbol Ketuhanan Yang Maha Esa.',
        'pancasila-bintang.jpg',
        1,
        'Bintang',
        'Pohon beringin',
        'Rantai',
        'Padi dan kapas',
        '',
        1,
    ])

    header_fill = PatternFill(fill_type='solid', fgColor='1D4ED8')
    header_font = Font(color='FFFFFF', bold=True)
    note_fill = PatternFill(fill_type='solid', fgColor='EFF6FF')

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for cell in sheet[2]:
        cell.fill = note_fill
        cell.alignment = Alignment(vertical='top', wrap_text=True)

    column_widths = {
        'A': 20,
        'B': 20,
        'C': 28,
        'D': 52,
        'E': 52,
        'F': 24,
        'G': 10,
        'H': 22,
        'I': 22,
        'J': 22,
        'K': 22,
        'L': 22,
        'M': 16,
    }
    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = 'A2'

    guide_sheet = workbook.create_sheet(title='Instructions')
    instructions = [
        ('Field', 'Keterangan'),
        ('parent_category', 'Opsional. Nama kategori utama, misalnya CPNS atau BUMN.'),
        ('category', 'Nama kategori soal.'),
        ('category_description', 'Opsional. Deskripsi kategori.'),
        ('question', 'Teks soal yang akan ditampilkan.'),
        ('explanation', 'Opsional. Pembahasan atau penjelasan jawaban yang akan ditampilkan di halaman hasil.'),
        ('image', 'Opsional. Nama file gambar. Isi kolom ini hanya jika upload via ZIP bersama file gambar.'),
        ('order', 'Urutan soal dalam kategori.'),
        ('choice_1 sampai choice_5', 'Pilihan jawaban. Minimal isi dua pilihan, maksimal lima.'),
        ('correct_answer', 'Nomor jawaban yang benar, isi 1 sampai 5 sesuai choice yang dipakai.'),
        ('Catatan', 'Untuk soal bergambar, upload file Excel ini di dalam ZIP bersama semua gambar yang dipakai.'),
    ]
    for row in instructions:
        guide_sheet.append(row)

    for cell in guide_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in guide_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    for column_index, width in enumerate((26, 90), start=1):
        guide_sheet.column_dimensions[get_column_letter(column_index)].width = width

    guide_sheet.freeze_panes = 'A2'

    return workbook


def is_blank_question_row(row):
    return not any(str(value).strip() for value in row.values())


def resolve_import_category(row, row_number):
    parent_category = None
    parent_category_name = row.get('parent_category', '')
    if parent_category_name:
        parent_category, _ = Category.objects.get_or_create(name=parent_category_name, parent=None)

    category_name = row.get('category', '')
    if not category_name:
        raise ValueError(f'Baris {row_number}: kolom "category" wajib diisi.')

    category = Category.objects.filter(name=category_name, parent=parent_category).first()
    if category:
        if not category.description and row.get('category_description', ''):
            category.description = row.get('category_description', '')
            category.save(update_fields=['description'])
        return category

    conflicting_category_exists = Category.objects.filter(name=category_name).exclude(
        parent=parent_category
    ).exists()
    if conflicting_category_exists:
        raise ValueError(
            f'Baris {row_number}: kategori "{category_name}" sudah ada pada struktur lain. '
            'Gunakan nama unik atau rapikan kategori yang ada terlebih dahulu.'
        )

    return Category.objects.create(
        name=category_name,
        parent=parent_category,
        description=row.get('category_description', ''),
    )


def validate_import_choices(row, row_number):
    choices = []
    for index in range(1, 6):
        choice_text = row.get(f'choice_{index}', '')
        if choice_text:
            choices.append((index, choice_text))

    if len(choices) < 2:
        raise ValueError(f'Baris {row_number}: minimal harus ada 2 pilihan jawaban.')

    correct_answer_str = row.get('correct_answer', '')
    if not correct_answer_str.isdigit():
        raise ValueError(f'Baris {row_number}: kolom "correct_answer" harus berisi angka 1-5.')

    correct_answer = int(correct_answer_str)
    valid_indexes = {index for index, _ in choices}
    if correct_answer not in valid_indexes:
        raise ValueError(
            f'Baris {row_number}: "correct_answer" harus menunjuk ke pilihan yang benar-benar terisi.'
        )

    return choices, correct_answer


def import_question_rows(rows, image_map):
    created_count = 0

    with transaction.atomic():
        for row_number, row in enumerate(rows, start=2):
            if is_blank_question_row(row):
                continue

            category = resolve_import_category(row, row_number)

            question_text = row.get('question', '')
            if not question_text:
                raise ValueError(f'Baris {row_number}: kolom "question" wajib diisi.')

            choices, correct_answer = validate_import_choices(row, row_number)

            try:
                order_val = int(row.get('order', 0))
            except ValueError:
                order_val = 0

            question = Question.objects.create(
                category=category,
                text=question_text,
                explanation=row.get('explanation', ''),
                order=order_val
            )

            image_filename = row.get('image', '')
            if image_filename:
                image_path = image_map.get(image_filename)
                if not image_path:
                    raise ValueError(
                        f'Baris {row_number}: file gambar "{image_filename}" tidak ditemukan di upload.'
                    )
                with open(image_path, 'rb') as image_file:
                    question.image.save(image_filename, File(image_file), save=True)

            for index, choice_text in choices:
                Choice.objects.create(
                    question=question,
                    text=choice_text,
                    is_correct=index == correct_answer,
                )

            created_count += 1

    return created_count


def build_admin_dashboard_stats():
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    pending_users = User.objects.filter(is_active=False).count()
    total_attempts = QuizAttempt.objects.count()
    attempts_today = QuizAttempt.objects.filter(completed_at__date=timezone.localdate()).count()
    total_categories = Category.objects.count()
    premium_categories = Category.objects.filter(is_premium=True).count()
    total_questions = Question.objects.count()
    total_choices = Choice.objects.count()

    attempts = QuizAttempt.objects.all()
    answered_sum = attempts.aggregate(total_score=Sum('score'), total_questions=Sum('total_questions'))
    total_score = answered_sum['total_score'] or 0
    total_question_count = answered_sum['total_questions'] or 0
    average_accuracy = round((total_score / total_question_count) * 100, 1) if total_question_count else 0.0

    top_categories = list(
        Category.objects.select_related('parent').annotate(attempt_count=Count('quizattempt'))
        .order_by('-attempt_count', 'name')[:5]
    )
    recent_attempts = list(
        QuizAttempt.objects.select_related('user', 'category__parent')
        .order_by('-completed_at')[:8]
    )

    return {
        'admin_stats': {
            'total_users': total_users,
            'active_users': active_users,
            'pending_users': pending_users,
            'total_attempts': total_attempts,
            'attempts_today': attempts_today,
            'total_categories': total_categories,
            'premium_categories': premium_categories,
            'total_questions': total_questions,
            'total_choices': total_choices,
            'average_accuracy': average_accuracy,
            'top_categories': top_categories,
            'recent_attempts': recent_attempts,
        }
    }

class ChoiceInline(admin.TabularInline):
    model = Choice
    extra = 5  # Changed from 4 to 5


class CategoryChoiceField(forms.ModelChoiceField):
    def label_from_instance(self, obj):
        return obj.full_name


class CategoryAdminForm(forms.ModelForm):
    parent = CategoryChoiceField(
        queryset=Category.objects.filter(parent__isnull=True).order_by('name'),
        required=False,
        help_text='Pilih kategori utama. Kosongkan jika kategori ini adalah kategori utama.',
    )

    class Meta:
        model = Category
        fields = '__all__'


class QuestionAdminForm(forms.ModelForm):
    category = CategoryChoiceField(
        queryset=Category.objects.select_related('parent').order_by('parent__name', 'name'),
    )

    class Meta:
        model = Question
        fields = '__all__'


class QuestionAdmin(admin.ModelAdmin):
    form = QuestionAdminForm
    inlines = [ChoiceInline]
    list_display = (
        'order',
        'category_full_name',
        'short_text',
        'has_image',
        'has_explanation',
    )
    list_filter = ('category__parent', 'category')
    search_fields = ('text', 'explanation', 'category__name', 'category__parent__name')
    ordering = ('category__parent__name', 'category__name', 'order', 'id')
    autocomplete_fields = ('category',)

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('category__parent')

    def category_full_name(self, obj):
        return obj.category.full_name
    category_full_name.short_description = 'Kategori'
    category_full_name.admin_order_field = 'category__name'

    def short_text(self, obj):
        if len(obj.text) <= 90:
            return obj.text
        return f'{obj.text[:87]}...'
    short_text.short_description = 'Soal'

    def has_image(self, obj):
        return format_html(
            '<span style="color: {};">{}</span>',
            '#059669' if obj.image else '#9CA3AF',
            'Ya' if obj.image else 'Tidak',
        )
    has_image.short_description = 'Gambar'

    def has_explanation(self, obj):
        return format_html(
            '<span style="color: {};">{}</span>',
            '#2563EB' if obj.explanation else '#9CA3AF',
            'Ada' if obj.explanation else 'Tidak',
        )
    has_explanation.short_description = 'Pembahasan'

class QuizAttemptAdmin(admin.ModelAdmin):
    list_display = ('user', 'category', 'score', 'total_questions', 'completed_at')
    list_filter = ('category', 'user', 'completed_at')
    readonly_fields = ('user', 'category', 'score', 'total_questions', 'completed_at', 'answers')

class SubscriptionAdmin(admin.ModelAdmin):
    list_display = (
        'user',
        'package',
        'max_attempts_per_quiz',
        'max_categories',
        'prefers_timer',
        'preferred_timer_minutes',
        'expires_at',
    )
    list_filter = ('package',)
    search_fields = ('user__username',)

class CategoryAdmin(admin.ModelAdmin):
    form = CategoryAdminForm
    list_display = (
        'full_name_display',
        'category_type',
        'is_premium',
        'question_count',
        'subcategory_count',
    )
    list_filter = ('is_premium', 'parent')
    search_fields = ('name', 'parent__name', 'description')
    ordering = ('parent__name', 'name', 'id')
    autocomplete_fields = ('parent',)
    fields = ('parent', 'name', 'description', 'image', 'is_premium')

    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .select_related('parent')
            .annotate(
                question_total=Count('questions', distinct=True),
                subcategory_total=Count('subcategories', distinct=True),
            )
        )

    def full_name_display(self, obj):
        if obj.parent_id:
            return format_html(
                '<strong>{}</strong><br><span style="color:#6B7280;">Subkategori dari {}</span>',
                obj.full_name,
                obj.parent.name,
            )
        return format_html(
            '<strong>{}</strong><br><span style="color:#6B7280;">Kategori utama</span>',
            obj.name,
        )
    full_name_display.short_description = 'Nama Kategori'
    full_name_display.admin_order_field = 'name'

    def category_type(self, obj):
        if obj.parent_id:
            return 'Subkategori'
        if obj.subcategory_total:
            return 'Induk'
        return 'Mandiri'
    category_type.short_description = 'Tipe'

    def question_count(self, obj):
        return obj.question_total
    question_count.short_description = 'Jumlah Soal'
    question_count.admin_order_field = 'question_total'

    def subcategory_count(self, obj):
        return obj.subcategory_total
    subcategory_count.short_description = 'Jumlah Subkategori'
    subcategory_count.admin_order_field = 'subcategory_total'
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-questions/', self.admin_site.admin_view(self.import_questions), name='import_questions'),
            path(
                'export-questions-template/',
                self.admin_site.admin_view(self.export_questions_template),
                name='export_questions_template',
            ),
        ]
        return custom_urls + urls

    def export_questions_template(self, request):
        workbook = build_question_template_workbook()
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="question-import-template.xlsx"'
        return response
    
    def import_questions(self, request):
        if request.method == 'POST':
            upload_file = request.FILES.get('csv_file')
            
            if not upload_file:
                messages.error(request, 'Silakan upload file CSV, XLSX, atau ZIP.')
                return redirect('..')
            
            is_zip = upload_file.name.lower().endswith('.zip')
            is_csv = upload_file.name.lower().endswith('.csv')
            is_xlsx = upload_file.name.lower().endswith('.xlsx')
            
            if not (is_zip or is_csv or is_xlsx):
                messages.error(request, 'File harus berformat CSV, XLSX, atau ZIP.')
                return redirect('..')
            
            temp_dir = tempfile.mkdtemp()
            try:
                question_file_path = None
                image_map = {}
                imported_from_zip = False
                
                if is_zip:
                    imported_from_zip = True
                    with zipfile.ZipFile(upload_file, 'r') as zip_ref:
                        zip_ref.extractall(temp_dir)
                    
                    for root, dirs, files in os.walk(temp_dir):
                        for file in files:
                            lower_name = file.lower()
                            if lower_name.endswith(('.csv', '.xlsx')) and not question_file_path:
                                question_file_path = os.path.join(root, file)
                            elif file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                                image_map[file] = os.path.join(root, file)
                                
                    if not question_file_path:
                        messages.error(request, 'Tidak ada file CSV atau XLSX di dalam arsip ZIP.')
                        return redirect('..')
                        
                else:
                    extension = '.xlsx' if is_xlsx else '.csv'
                    question_file_path = os.path.join(temp_dir, f'import{extension}')
                    with open(question_file_path, 'wb+') as destination:
                        for chunk in upload_file.chunks():
                            destination.write(chunk)
                
                rows, source_label = load_question_rows(question_file_path)
                created_count = import_question_rows(rows, image_map)
                messages.success(
                    request,
                    f'Berhasil mengimpor {created_count} soal'
                    f'{" beserta gambar" if imported_from_zip else ""} dari {source_label}!'
                )
                return redirect('..')
                
            except Exception as e:
                messages.error(request, f'Terjadi kesalahan saat import: {str(e)}')
                return redirect('..')
            finally:
                if os.path.exists(temp_dir):
                    shutil.rmtree(temp_dir)
        
        return render(request, 'admin/import_questions.html')

admin.site.register(Category, CategoryAdmin)
admin.site.register(Question, QuestionAdmin)
admin.site.register(QuizAttempt, QuizAttemptAdmin)
admin.site.register(Subscription, SubscriptionAdmin)

original_admin_index = admin.site.index


def custom_admin_index(request, extra_context=None):
    context = {}
    if extra_context:
        context.update(extra_context)
    context.update(build_admin_dashboard_stats())
    return original_admin_index(request, extra_context=context)


admin.site.index = custom_admin_index
admin.site.index_template = 'admin/dashboard_index.html'
admin.site.site_header = "Administrasi BelajarUji CPNS & BUMN"
admin.site.site_title = "BelajarUji Admin"
admin.site.index_title = "Panel Admin BelajarUji"
